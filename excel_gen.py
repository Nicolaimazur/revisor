"""
Excel template engine.

Brugeren uploader en .xlsx-skabelon med {{placeholder}}-variabler.
Platformen erstatter variablerne med data fra rapporten og returnerer
det udfyldte Excel-dokument.

Eksempel i skabelon-celle:  {{selskabsnavn}}
Output:                     ACME ApS
"""

import io
import re
from openpyxl import load_workbook


# ── Variable-mapping ──────────────────────────────────────────────────────────

def _build_vars(report_data: dict) -> dict[str, str]:
    """Byg {{{placeholder}}: tekst}-mapping fra rapport-data."""
    s01  = report_data.get("sektion_01")  or {}
    s02  = report_data.get("sektion_02")  or {}
    s03  = report_data.get("sektion_03")  or {}
    s04  = report_data.get("sektion_04")  or {}
    s07  = report_data.get("sektion_07")  or {}
    s09  = report_data.get("sektion_09")  or {}
    meta = report_data.get("meta")        or {}

    # ── Ledelse ───────────────────────────────────────────────────────────────
    direktion = s02.get("direktion") or []
    dir_lines = "\n".join(
        f"{d.get('navn','')}  ({d.get('rolle','')})"
        for d in direktion if d.get("navn")
    )
    bestyrelse = "\n".join(
        f"{d.get('navn','')}  ({d.get('rolle','')})"
        for d in direktion
        if any(k in d.get("rolle", "") for k in ["Bestyrelsesmedlem", "Bestyrelsesformand"])
    )

    # ── Ejere ─────────────────────────────────────────────────────────────────
    ejere = s02.get("reelle_ejere") or []
    ejer_lines = "\n".join(
        f"{e.get('navn','')}  {e.get('ejerandel','')}  ({e.get('besiddelse','')})"
        for e in ejere if e.get("navn")
    )

    # ── Finansielle tal ───────────────────────────────────────────────────────
    aar_kolonner = [str(a) for a in (s04.get("aar_kolonner") or []) if a]
    reg_rows     = s04.get("regnskabspost_tabel") or []

    def _fin(keyword: str, year_idx: int = 0) -> str:
        for r in reg_rows:
            if keyword.lower() in r.get("post", "").lower():
                vals = r.get("vaerdier") or []
                v = vals[year_idx] if year_idx < len(vals) else None
                return str(v) if v is not None else "—"
        return "—"

    def _yr(idx: int) -> str:
        return aar_kolonner[idx] if idx < len(aar_kolonner) else ""

    # ── Risikooversigt ────────────────────────────────────────────────────────
    risiko = {r.get("indikator", ""): r.get("vurdering", "")
              for r in (report_data.get("risiko_oversigt") or [])}

    raw = {
        # — Meta / datoid —
        "rapport_dato":            meta.get("rapport_dato", ""),

        # — Sektion 01: CVR-data —
        "selskabsnavn":            s01.get("selskabsnavn", ""),
        "cvr":                     meta.get("cvr", ""),
        "adresse":                 s01.get("adresse", ""),
        "stiftelsesdato":          s01.get("stiftelsesdato", ""),
        "selskabsform":            s01.get("selskabsform", ""),
        "branchekode":             s01.get("branchekode", ""),
        "formaal":                 s01.get("formaal", ""),
        "status":                  s01.get("status", ""),
        "website":                 s01.get("website", ""),
        "selskabsbeskrivelse":     s01.get("selskabsbeskrivelse", ""),
        "seneste_regnskabsaar":    s01.get("seneste_regnskabsaar") or _yr(0),
        "naeste_regnskabsfrist":   s01.get("naeste_regnskabsfrist", ""),
        "reklamebeskyttet":        s01.get("reklamebeskyttet", ""),

        # — Sektion 02: Ledelse & ejere —
        "direktion":               dir_lines,
        "bestyrelse":              bestyrelse,
        "ejerforhold":             ejer_lines,
        "ubo_note":                s02.get("ubo_note", ""),
        "pep_screening_note":      s02.get("pep_screening_note", ""),

        # — Sektion 03: Ejerstruktur —
        "organisationsstruktur":   s03.get("beskrivelse", ""),
        "revisor_anbefaling_s03":  s03.get("revisor_anbefaling", ""),

        # — Sektion 04: Økonomi —
        "seneste_aar":             _yr(0),
        "forrige_aar":             _yr(1),
        "aar_3":                   _yr(2),
        "aar_4":                   _yr(3),
        "aar_5":                   _yr(4),

        "bruttofortjeneste_seneste": _fin("bruttofortjeneste", 0),
        "bruttofortjeneste_forrige": _fin("bruttofortjeneste", 1),
        "bruttofortjeneste_aar3":    _fin("bruttofortjeneste", 2),

        "resultat_seneste":        _fin("resultat", 0),
        "resultat_forrige":        _fin("resultat", 1),
        "resultat_aar3":           _fin("resultat", 2),

        "egenkapital_seneste":     _fin("egenkapital", 0),
        "egenkapital_forrige":     _fin("egenkapital", 1),
        "egenkapital_aar3":        _fin("egenkapital", 2),

        "balance_seneste":         _fin("balance", 0),
        "balance_forrige":         _fin("balance", 1),
        "balance_aar3":            _fin("balance", 2),

        "ebit_seneste":            _fin("ebit", 0),
        "ebit_forrige":            _fin("ebit", 1),

        "finansiel_analyse":       s04.get("finansiel_analyse", ""),

        # — Sektion 07: Branche —
        "branche_karakter":        s07.get("branchekarakter", ""),
        "markedssituation":        s07.get("markedssituation", ""),
        "benchmarking_note":       s07.get("benchmarking_note", ""),

        # — Sektion 09: Risikovurdering —
        "samlet_vurdering":        s09.get("samlet_vurdering", ""),
        "risiko_niveau":           s09.get("risiko_niveau", "").upper(),
        "ansvarsfraskrivelse":     s09.get("ansvarsfraskrivelse", ""),
    }

    # Wrap keys med {{ }}
    return {f"{{{{{k}}}}}": (v or "") for k, v in raw.items()}


# ── Template-udfyldning ───────────────────────────────────────────────────────

def fill_template(template_bytes: bytes, report_data: dict) -> bytes:
    """
    Indlæs Excel-skabelon, erstat alle {{variabler}} med rapport-data,
    og returnér det udfyldte Excel-dokument som bytes.
    """
    var_map = _build_vars(report_data)
    wb = load_workbook(io.BytesIO(template_bytes))

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                val = cell.value
                changed = False
                for placeholder, replacement in var_map.items():
                    if placeholder in val:
                        val = val.replace(placeholder, replacement)
                        changed = True
                if changed:
                    cell.value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tilgængelige variabler (til visning i UI) ─────────────────────────────────

AVAILABLE_PLACEHOLDERS: list[dict] = [
    # Grunddata
    {"var": "{{selskabsnavn}}",         "beskrivelse": "Virksomhedens fulde navn"},
    {"var": "{{cvr}}",                  "beskrivelse": "CVR-nummer"},
    {"var": "{{adresse}}",              "beskrivelse": "Registreret adresse"},
    {"var": "{{stiftelsesdato}}",       "beskrivelse": "Stiftelsesdato"},
    {"var": "{{selskabsform}}",         "beskrivelse": "Selskabsform (ApS, A/S mv.)"},
    {"var": "{{branchekode}}",          "beskrivelse": "Branchekode og beskrivelse"},
    {"var": "{{formaal}}",              "beskrivelse": "Selskabets formål"},
    {"var": "{{status}}",               "beskrivelse": "CVR-status"},
    {"var": "{{website}}",              "beskrivelse": "Website"},
    {"var": "{{selskabsbeskrivelse}}", "beskrivelse": "AI-genereret selskabsbeskrivelse"},
    {"var": "{{seneste_regnskabsaar}}", "beskrivelse": "Seneste regnskabsår (YYYY)"},
    {"var": "{{naeste_regnskabsfrist}}","beskrivelse": "Næste indberetningsfrist"},
    {"var": "{{rapport_dato}}",         "beskrivelse": "Dato rapporten er genereret"},
    # Ledelse & ejere
    {"var": "{{direktion}}",            "beskrivelse": "Direktionsmedlemmer — én linje per person"},
    {"var": "{{bestyrelse}}",           "beskrivelse": "Bestyrelsesmedlemmer — én linje per person"},
    {"var": "{{ejerforhold}}",          "beskrivelse": "Reelle ejere med ejerandel — én linje per ejer"},
    {"var": "{{organisationsstruktur}}","beskrivelse": "Beskrivelse af koncern- og ejerstruktur"},
    # Økonomi (seneste år)
    {"var": "{{seneste_aar}}",          "beskrivelse": "Årstal — seneste regnskabsår"},
    {"var": "{{forrige_aar}}",          "beskrivelse": "Årstal — foregående år"},
    {"var": "{{bruttofortjeneste_seneste}}", "beskrivelse": "Bruttofortjeneste — seneste år (t.DKK)"},
    {"var": "{{bruttofortjeneste_forrige}}", "beskrivelse": "Bruttofortjeneste — foregående år"},
    {"var": "{{resultat_seneste}}",     "beskrivelse": "Årets resultat — seneste år (t.DKK)"},
    {"var": "{{resultat_forrige}}",     "beskrivelse": "Årets resultat — foregående år"},
    {"var": "{{egenkapital_seneste}}",  "beskrivelse": "Egenkapital — seneste år (t.DKK)"},
    {"var": "{{egenkapital_forrige}}",  "beskrivelse": "Egenkapital — foregående år"},
    {"var": "{{balance_seneste}}",      "beskrivelse": "Balance (aktiver) — seneste år (t.DKK)"},
    {"var": "{{finansiel_analyse}}",    "beskrivelse": "AI-genereret finansiel analyse (lang tekst)"},
    # Risikovurdering
    {"var": "{{samlet_vurdering}}",     "beskrivelse": "Samlet risikovurdering (tekst)"},
    {"var": "{{risiko_niveau}}",        "beskrivelse": "Risikoniveau: LAV / MIDDEL / HØJ"},
    {"var": "{{markedssituation}}",     "beskrivelse": "Branchebeskrivelse og markedssituation"},
]
