"""
AI-powered template analyzer.

When a user uploads an Excel or PDF template the module:
1. Extracts the document structure (grid text with [TOM:COORD] markers for empty cells)
2. Asks Claude to identify which [TOM:COORD] cells should contain which report field
3. Validates the result — discards any cell Claude pointed at that already had content
4. Stores the analysis on disk

At download time fill_from_mapping() uses the stored analysis.
"""

import io
import json
import re
from pathlib import Path

import anthropic


# ── Field catalogue ───────────────────────────────────────────────────────────

FIELDS = {
    "selskabsnavn":              "Virksomhedens fulde navn",
    "cvr":                       "CVR-nummer (8 cifre)",
    "adresse":                   "Registreret adresse",
    "stiftelsesdato":            "Stiftelsesdato",
    "selskabsform":              "Selskabsform (ApS, A/S osv.)",
    "branchekode":               "Branchekode og -beskrivelse",
    "formaal":                   "Selskabets formål ifølge vedtægterne",
    "status":                    "CVR-status (NORMAL / UNDER KONKURS osv.)",
    "website":                   "Virksomhedens hjemmeside",
    "selskabsbeskrivelse":       "Kort AI-genereret beskrivelse af virksomheden",
    "rapport_dato":              "Dato rapporten er genereret",
    "seneste_regnskabsaar":      "Seneste regnskabsår (YYYY)",
    "naeste_regnskabsfrist":     "Næste indberetningsfrist",
    "direktion":                 "Direktionsmedlemmer, én linje per person",
    "bestyrelse":                "Bestyrelsesmedlemmer, én linje per person",
    "ejerforhold":               "Reelle ejere med ejerandel, én linje per ejer",
    "organisationsstruktur":     "Beskrivelse af koncern- og ejerstruktur",
    "ubo_note":                  "Note om ultimate beneficial owners (UBO)",
    "pep_screening_note":        "Note om PEP / politisk eksponerede personer",
    "seneste_aar":               "Årstal — seneste regnskabsår",
    "forrige_aar":               "Årstal — foregående år",
    "bruttofortjeneste_seneste": "Bruttofortjeneste seneste år (t.DKK)",
    "bruttofortjeneste_forrige": "Bruttofortjeneste foregående år (t.DKK)",
    "bruttofortjeneste_aar3":    "Bruttofortjeneste 3 år siden (t.DKK)",
    "resultat_seneste":          "Årets resultat seneste år (t.DKK)",
    "resultat_forrige":          "Årets resultat foregående år (t.DKK)",
    "resultat_aar3":             "Årets resultat 3 år siden (t.DKK)",
    "egenkapital_seneste":       "Egenkapital seneste år (t.DKK)",
    "egenkapital_forrige":       "Egenkapital foregående år (t.DKK)",
    "egenkapital_aar3":          "Egenkapital 3 år siden (t.DKK)",
    "balance_seneste":           "Balance (aktiver) seneste år (t.DKK)",
    "balance_forrige":           "Balance foregående år (t.DKK)",
    "balance_aar3":              "Balance 3 år siden (t.DKK)",
    "ebit_seneste":              "EBIT seneste år (t.DKK)",
    "ebit_forrige":              "EBIT foregående år (t.DKK)",
    "finansiel_analyse":         "AI-genereret finansiel analyse (lang tekst)",
    "branche_karakter":          "Branchekarakter",
    "markedssituation":          "Branchebeskrivelse og markedssituation",
    "samlet_vurdering":          "Samlet risikovurdering (tekst)",
    "risiko_niveau":             "Risikoniveau: LAV / MIDDEL / HØJ",
    "ansvarsfraskrivelse":       "Ansvarsfraskrivelse",
}

_FIELDS_HINT = "\n".join(f"  {k}: {v}" for k, v in FIELDS.items())


# ── Structure extraction ──────────────────────────────────────────────────────

def _extract_xlsx(xlsx_bytes: bytes):
    """
    Returns (grid_text, empty_cells_set).

    grid_text  — human-readable layout where empty cells show as [TOM:COORD].
                 Claude uses this to understand label→value relationships.
    empty_cells — set of all cell coordinates that were truly empty.
                 Used to validate Claude's output (it must only map empty cells).
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    lines = []
    empty_cells = set()   # {(sheet_name, coord), ...}

    for ws in wb.worksheets:
        lines.append(f"\nArk: {ws.title}")
        max_row = min(ws.max_row or 0, 200)
        # Always scan at least 3 cols beyond last filled, min 8
        max_col = min(max((ws.max_column or 1) + 3, 8), 20)

        for r in range(1, max_row + 1):
            parts = []
            has_content = False
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                col  = get_column_letter(c)
                coord = f"{col}{r}"
                v = cell.value
                if v is not None and str(v).strip():
                    parts.append(f'{col}="{str(v).strip()[:80]}"')
                    has_content = True
                else:
                    parts.append(f"{col}=[TOM:{coord}]")
                    empty_cells.add((ws.title, coord))
            if has_content:
                lines.append(f"  Række {r}: " + "  ".join(parts))

    return "\n".join(lines), empty_cells


def _extract_pdf(pdf_bytes: bytes) -> list:
    """Return [{page, text, tables}] for each page (max 10)."""
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("Installer pdfplumber: pip install pdfplumber")

    pages = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for i, page in enumerate(pdf.pages[:10]):
            raw_tables = page.extract_tables() or []
            tables = []
            for tbl in raw_tables:
                clean = [[str(c or "").strip() for c in row] for row in tbl if any(c for c in row)]
                if clean:
                    tables.append(clean[:20])
            pages.append({
                "page": i + 1,
                "text": (page.extract_text() or "")[:4000],
                "tables": tables[:4],
            })
    return pages


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_json(text: str) -> dict:
    text = text.strip()
    m = re.search(r"```(?:json)?\s*([\s\S]+?)```", text)
    if m:
        text = m.group(1).strip()
    start = text.find("{")
    if start != -1:
        text = text[start:]
    return json.loads(text)


def _validate_xlsx_mapping(raw_mapping: dict, empty_cells: set) -> dict:
    """
    Discard any cell Claude mapped that was NOT empty in the original template.
    This prevents Claude from accidentally overwriting label cells.
    Also discard entries whose field name isn't in our FIELDS catalogue.
    """
    clean = {}
    for sheet, cell_map in raw_mapping.items():
        clean_sheet = {}
        for coord, field in cell_map.items():
            if field not in FIELDS:
                continue  # unknown field
            if (sheet, coord) in empty_cells:
                clean_sheet[coord] = field
            else:
                # Claude pointed at a non-empty cell — skip it
                pass
        if clean_sheet:
            clean[sheet] = clean_sheet
    return clean


# ── Claude analysis ───────────────────────────────────────────────────────────

async def _analyze_xlsx_structure(grid: str, empty_cells: set) -> dict:
    if len(grid) > 14000:
        grid = grid[:14000] + "\n…(afkortet)"

    # Build a set-description for the prompt so Claude knows the rule
    prompt = f"""Du er ekspert i Excel-dokumentstruktur og KYC/due diligence for revisorer.

Her er layoutet af en Excel KYC-skabelon.
Celler markeret [TOM:COORD] er TOMME — det er dem der skal udfyldes.
Celler med tekst er labels/overskrifter — de må ALDRIG overskrives.

{grid}

Tilgængelige datafelter:
{_FIELDS_HINT}

REGLER:
1. Du må KUN returnere koordinater der optræder som [TOM:COORD] i layoutet ovenfor.
   Returner ALDRIG en celle der allerede har indhold.
2. Identificer label-cellen (til venstre, over, eller i nærheden) og brug den til
   at afgøre hvilket felt der passer til den tilstødende tomme celle.
3. Vær UDTØMMENDE — find alle felter du kan identificere. Det er bedre at have
   for mange end for få.
4. Brug fuzzy-match på labels: "Firmanavn" → selskabsnavn, "Reg.dato" → stiftelsesdato,
   "Org.nr" / "CVR" → cvr, "Direktør" → direktion, "Overskud" → resultat_seneste, osv.

Returner KUN gyldigt JSON — ingen forklaring:
{{
  "Arknavn": {{
    "B3": "selskabsnavn",
    "B4": "cvr",
    "B5": "adresse"
  }}
}}"""

    client = anthropic.AsyncAnthropic()
    msg = await client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = _parse_json(msg.content[0].text)
    return _validate_xlsx_mapping(raw, empty_cells)


async def _analyze_pdf_structure(pages: list) -> dict:
    pages_txt = json.dumps(pages, ensure_ascii=False, indent=2)
    if len(pages_txt) > 14000:
        pages_txt = pages_txt[:14000] + "\n…(afkortet)"

    prompt = f"""Du analyserer et PDF KYC/due diligence skabelon-dokument brugt af revisorer.

Dokumentets indhold:
{pages_txt}

Tilgængelige datafelter:
{_FIELDS_HINT}

OPGAVE — vær UDTØMMENDE:
1. Map alle labels i dokumentet til de rigtige felter.
   Brug bredt match: "Firmanavn"/"Selskab" → selskabsnavn, "CVR"/"Org.nr" → cvr,
   "Stiftet"/"Etableret" → stiftelsesdato, "Direktør"/"Ledelse" → direktion,
   "Ejer"/"Aktionær" → ejerforhold, "Resultat"/"Overskud" → resultat_seneste,
   "Egenkapital" → egenkapital_seneste, "Balance"/"Aktiver" → balance_seneste,
   "Risiko" → risiko_niveau eller samlet_vurdering, osv.
2. Byg sektioner der matcher dokumentets struktur, og inkluder ALLE relevante felter
   i hver sektion — selv om de ikke eksplicit er navngivet i dokumentet.

Returner KUN gyldigt JSON:
{{
  "felt_map": [
    {{"label": "Virksomhedsnavn", "field": "selskabsnavn"}},
    {{"label": "CVR-nr.", "field": "cvr"}}
  ],
  "sections": [
    {{"titel": "Stamoplysninger", "felter": ["selskabsnavn","cvr","adresse","stiftelsesdato","selskabsform","branchekode","status","website"]}},
    {{"titel": "Ledelse & ejere",  "felter": ["direktion","bestyrelse","ejerforhold","ubo_note","pep_screening_note"]}},
    {{"titel": "Økonomi",          "felter": ["seneste_aar","resultat_seneste","resultat_forrige","egenkapital_seneste","egenkapital_forrige","balance_seneste","bruttofortjeneste_seneste","finansiel_analyse"]}},
    {{"titel": "Risikovurdering",  "felter": ["risiko_niveau","samlet_vurdering","markedssituation","ansvarsfraskrivelse"]}}
  ]
}}"""

    client = anthropic.AsyncAnthropic()
    msg = await client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2500,
        messages=[{"role": "user", "content": prompt}],
    )
    return _parse_json(msg.content[0].text)


# ── Public: analyze ───────────────────────────────────────────────────────────

async def analyze_template(file_bytes: bytes, filename: str) -> dict:
    """
    Analyze an uploaded template. Returns analysis dict to persist on disk.
      xlsx → {"type": "xlsx", "mapping": {sheet: {cell: field}}}
      pdf  → {"type": "pdf",  "mapping": {felt_map:[...], sections:[...]}}
    """
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        grid, empty_cells = _extract_xlsx(file_bytes)
        mapping = await _analyze_xlsx_structure(grid, empty_cells)
        return {"type": "xlsx", "mapping": mapping}
    elif ext == ".pdf":
        pages = _extract_pdf(file_bytes)
        mapping = await _analyze_pdf_structure(pages)
        return {"type": "pdf", "mapping": mapping}
    else:
        raise ValueError(f"Understøtter kun .xlsx og .pdf, ikke '{ext}'")


# ── Public: preview ───────────────────────────────────────────────────────────

def preview_mapping(analysis: dict, report_data: dict) -> list:
    """
    Returns a list of {cell, field, label, value} for UI display —
    so the user can verify the mapping before downloading.
    """
    from excel_gen import _build_vars
    var_map = _build_vars(report_data)
    vals = {k[2:-2]: v for k, v in var_map.items()}

    tpl_type = analysis.get("type", "xlsx")
    mapping  = analysis.get("mapping", {})
    rows = []

    if tpl_type == "xlsx":
        for sheet, cell_map in mapping.items():
            for coord, field in sorted(cell_map.items()):
                rows.append({
                    "cell":  f"{sheet}!{coord}",
                    "field": field,
                    "label": FIELDS.get(field, field),
                    "value": vals.get(field, ""),
                })
    else:  # pdf
        felt_map = {item["field"]: item["label"] for item in mapping.get("felt_map", [])}
        for field, label in felt_map.items():
            rows.append({
                "cell":  "—",
                "field": field,
                "label": label,
                "value": vals.get(field, ""),
            })
    return rows


# ── Public: fill ──────────────────────────────────────────────────────────────

def fill_from_mapping(template_bytes: bytes, analysis: dict, report_data: dict) -> bytes:
    """Fill template using stored analysis. Returns xlsx bytes."""
    from excel_gen import _build_vars
    from openpyxl import load_workbook

    var_map = _build_vars(report_data)
    vals = {k[2:-2]: v for k, v in var_map.items()}

    tpl_type = analysis.get("type", "xlsx")
    mapping  = analysis.get("mapping", {})

    if tpl_type == "xlsx":
        wb = load_workbook(io.BytesIO(template_bytes))
        for sheet_name, cell_map in mapping.items():
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            for cell_addr, field in cell_map.items():
                value = vals.get(field, "")
                if value:          # only write non-empty values
                    try:
                        ws[cell_addr] = value
                    except Exception:
                        pass
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    elif tpl_type == "pdf":
        return _build_xlsx_from_pdf_analysis(mapping, vals)

    else:
        raise ValueError(f"Ukendt skabelontype: {tpl_type}")


def _build_xlsx_from_pdf_analysis(mapping: dict, vals: dict) -> bytes:
    """Generate a structured Excel from PDF analysis + report data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "KYC Rapport"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 52

    felt_map = {item["field"]: item["label"] for item in mapping.get("felt_map", [])}
    sections = mapping.get("sections", [])
    if not sections:
        sections = [{"titel": "Stamoplysninger", "felter": list(felt_map.keys())}]

    row = 1
    for section in sections:
        hdr = ws.cell(row=row, column=1, value=section.get("titel", ""))
        hdr.font      = Font(bold=True, size=11, color="FFFFFF")
        hdr.fill      = PatternFill("solid", fgColor="2B3437")
        hdr.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 22
        row += 1

        for field in section.get("felter", []):
            lbl = felt_map.get(field, FIELDS.get(field, field))
            val = vals.get(field, "")

            lbl_cell = ws.cell(row=row, column=1, value=lbl + ":")
            lbl_cell.font      = Font(bold=True, size=10)
            lbl_cell.fill      = PatternFill("solid", fgColor="EDF1F3")
            lbl_cell.alignment = Alignment(vertical="top", wrap_text=True)

            val_cell = ws.cell(row=row, column=2, value=val)
            val_cell.font      = Font(size=10)
            val_cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.row_dimensions[row].height = 40 if len(str(val)) > 80 else 18
            row += 1

        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
